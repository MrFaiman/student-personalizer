"""
In-memory school data generator for debug/seeding purposes.
Adapted from data/generate_school_data.py to return BytesIO instead of writing files.
"""

import random
from io import BytesIO

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# names

MALE_FIRST = [
    "אביב", "אדם", "אהרון", "אור", "אורי", "אושר", "איתי", "איתן",
    "אלון", "אמיר", "ארי", "אריאל", "בן", "בנימין", "גל", "גיא",
    "דביר", "דוד", "דור", "דניאל", "האני", "הילל", "טל", "יובל",
    "יונתן", "יוסף", "יחיאל", "יניב", "ירון", "כפיר", "לביא", "לי",
    "ליאב", "ליאור", "ליאם", "מיכאל", "מתן", "נדב", "נוה", "נועם",
    "נחום", "ניב", "ניר", "עדי", "עומר", "עמית", "עמנואל", "פלא",
    "צח", "רועי", "רז", "רן", "רני", "שחר", "שי", "תום",
]

FEMALE_FIRST = [
    "אביבה", "אדוה", "אור", "אורי", "אורית", "אינה", "אמה", "אנה",
    "אסתר", "בת‑אל", "גל", "גלית", "דנה", "הדס", "הילה", "טל",
    "יעל", "ירדן", "כרמל", "לי", "ליאל", "ליאת", "לימור", "מאיה",
    "מיה", "מיכל", "מירי", "נועה", "נטע", "נילי", "עדי", "עינב",
    "עמית", "ענת", "פלג", "קרן", "רונית", "רחל", "רינת", "שי",
    "שיר", "שירה", "שרה", "תהל", "תמר",
]

LAST_NAMES = [
    "אברהם", "אדלר", "אהרוני", "אוחיון", "אזולאי", "אחדות", "אטיאס",
    "אלבז", "אלון", "אלחנני", "אליהו", "אלמוג", "אמסלם", "אסרף",
    "אפרים", "בן‑דוד", "בן‑יוסף", "בן‑לולו", "בן‑שלמה", "ביטון",
    "בירנבוים", "בן‑חיים", "גבאי", "גולן", "גורן", "גטניו", "דהן",
    "דוד", "הגר", "הרוש", "וקנין", "זוהר", "זכריה", "חדד", "חיון",
    "חיים", "טל", "יוסף", "יחיאלי", "יעקב", "כהן", "כץ", "לוי",
    "לומברוזו", "מזרחי", "מלכה", "מנחם", "מצרי", "נחמני", "סבג",
    "סגל", "ספיר", "עמר", "פרדו", "פרץ", "צדוק", "קדוש", "קורן",
    "ראובן", "רוזן", "שאבי", "שפירא", "שימאי", "שלמה", "שמיר",
]

SUBJECTS = {
    "מתמטיקה":       {"num_grades": 12, "mean": 68, "std": 22},
    "אנגלית":        {"num_grades": 10, "mean": 70, "std": 20},
    "מדעי המחשב":    {"num_grades":  8, "mean": 72, "std": 18},
    "פיזיקה":        {"num_grades":  8, "mean": 64, "std": 23},
    "ספרות":         {"num_grades":  6, "mean": 74, "std": 16},
    "היסטוריה":      {"num_grades":  6, "mean": 71, "std": 17},
    "אזרחות":        {"num_grades":  5, "mean": 76, "std": 15},
    "סייבר":         {"num_grades":  7, "mean": 73, "std": 19},
    "ביולוגיה":      {"num_grades":  7, "mean": 67, "std": 21},
    "חינוך גופני":   {"num_grades":  4, "mean": 82, "std": 10},
}

EVENT_COLS = [
    ("שיעורים שדווחו",              450,  60),
    ("חיסור",                       30,   25),
    ("חיסור (מוצדק)",               10,   10),
    ("איחור",                        4,    5),
    ("איחור (מוצדק)",                2,    3),
    ("הפרעה",                        2,    3),
    ("אי כניסה לשיעור",              1,    2),
    ("אי כניסה לשיעור (מוצדק)",      0,    1),
    ("תלבושת",                       1,    2),
    ("אי הכנת ש.ב",                  3,    4),
    ("חיזוק חיובי",                  4,    6),
    ("חיזוק חיובי כיתתי",            1,    2),
    ("אי הבאת ציוד",                 2,    3),
    ("אי ביצוע מטלות בכיתה",         1,    2),
    ('שימוש בנייד בשטח ביה"ס',       1,    2),
    ("היעדרות בפרטני (מוצדק)",       0,    1),
    ("נוכחות בפרטני",               5,    8),
]

# style

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHDR_FILL = PatternFill("solid", start_color="2E75B6")
SUBHDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DATA_FONT   = Font(name="Arial", size=9)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")
CENTER      = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="BDD7EE"),
    right=Side(style="thin", color="BDD7EE"),
    top=Side(style="thin", color="BDD7EE"),
    bottom=Side(style="thin", color="BDD7EE"),
)


def _style_header(cell, sub=False):
    cell.fill = SUBHDR_FILL if sub else HEADER_FILL
    cell.font = SUBHDR_FONT if sub else HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _style_data(cell, row_idx):
    cell.font = DATA_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    if row_idx % 2 == 0:
        cell.fill = ALT_FILL


def _grade_color(cell, value):
    if value is None:
        return
    if value >= 80:
        cell.fill = PatternFill("solid", start_color="C6EFCE")
        cell.font = Font(name="Arial", size=9, color="276221")
    elif value >= 55:
        cell.fill = PatternFill("solid", start_color="FFEB9C")
        cell.font = Font(name="Arial", size=9, color="9C5700")
    else:
        cell.fill = PatternFill("solid", start_color="FFC7CE")
        cell.font = Font(name="Arial", size=9, color="9C0006")
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def _auto_col_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 3, 30))


# generation

def _random_name(rng: random.Random, gender="M"):
    first = rng.choice(MALE_FIRST if gender == "M" else FEMALE_FIRST)
    last = rng.choice(LAST_NAMES)
    return f"{first} {last}"


def _generate_students(rng: random.Random, n=120):
    students = []
    used_tz: set[str] = set()
    for i in range(1, n + 1):
        gender = rng.choice(["M", "F"])
        name = _random_name(rng, gender)
        grade = "י"
        cls = rng.randint(1, 4)
        while True:
            tz = str(rng.randint(100_000_000, 999_999_999))
            if tz not in used_tz:
                used_tz.add(tz)
                break
        students.append((i, tz, name, grade, cls))
    return students


def _generate_teachers(rng: random.Random, subjects: dict) -> dict[str, str]:
    teachers: dict[str, str] = {}
    used: set[str] = set()
    for subj in subjects:
        while True:
            name = _random_name(rng, rng.choice(["M", "F"]))
            if name not in used:
                used.add(name)
                teachers[subj] = name
                break
    return teachers


def _student_base_ability(np_rng: np.random.Generator, n=120):
    return np.clip(np_rng.normal(0.5, 0.18, n), 0.1, 0.9)


def _generate_grade(np_rng: np.random.Generator, ability, mean, std, period_drift=0):
    personal_mean = mean + (ability - 0.5) * 40 + period_drift
    g = np_rng.normal(personal_mean, std * 0.6)
    return int(np.clip(round(g), 0, 100))


def _grades_for_student(rng: random.Random, np_rng: np.random.Generator, ability, subject_cfg, null_prob=0.10, period_drift=0):
    n = subject_cfg["num_grades"]
    grades = []
    for _ in range(n):
        if rng.random() < null_prob:
            grades.append(None)
        else:
            grades.append(_generate_grade(np_rng, ability, subject_cfg["mean"], subject_cfg["std"], period_drift))
    return grades


def _generate_events_for_student(rng: random.Random, np_rng: np.random.Generator, ability, period_drift=0, null_prob=0.35):
    row = {}
    behaviour_factor = 1.5 - ability
    for col, mean, std in EVENT_COLS:
        if col == "שיעורים שדווחו":
            val = int(np.clip(round(np_rng.normal(mean + period_drift * 2, std)), 300, 600))
            row[col] = val
        elif col in ("חיזוק חיובי", "חיזוק חיובי כיתתי", "נוכחות בפרטני"):
            adj_mean = mean * (0.5 + ability)
            val = max(0, int(round(np_rng.normal(adj_mean, std))))
            row[col] = val if val > 0 else (None if rng.random() < null_prob else 0)
        else:
            adj_mean = mean * behaviour_factor
            val = max(0, int(round(np_rng.normal(adj_mean, std))))
            row[col] = val if val > 0 else (None if rng.random() < null_prob else 0)
    return row


def _build_periods(years: list[int], quarters: int = 4) -> list[tuple[str, str, int]]:
    QUARTER_DRIFT = {1: -1, 2: 2, 3: 3, 4: 1}
    periods = []
    for y_idx, year in enumerate(years):
        year_label = str(year)
        year_bonus = round(y_idx * 0.5)
        for q in range(1, quarters + 1):
            drift = QUARTER_DRIFT[q] + year_bonus
            periods.append((f"Q{q}", year_label, drift))
    return periods


# build excel

def _build_grades_excel_bytes(rng, np_rng, students, subjects, teachers, abilities, period, year, period_drift) -> bytes:
    wb = Workbook()
    ws = wb.active
    grade_level = students[0][3]
    ws.title = f"{grade_level} - ציונים שוטפים - ממוצע"

    headers = ["מס'", "ת.ז", "שם התלמיד", "שכבה", "כיתה"]
    for subj, cfg in subjects.items():
        col_name = f"{subj} - {teachers[subj]}"
        headers.extend([col_name] * cfg["num_grades"])
    headers.append("ממוצע")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        _style_header(cell)
    ws.row_dimensions[1].height = 20

    for row_i, (num, tz, name, grade_level, cls) in enumerate(students, 2):
        ability = abilities[row_i - 2]
        for c, val in enumerate([num, tz, name, grade_level, cls], 1):
            _style_data(ws.cell(row_i, c, val), row_i)

        all_valid = []
        col = 6
        for subj, cfg in subjects.items():
            grades = _grades_for_student(rng, np_rng, ability, cfg, period_drift=period_drift)
            for g in grades:
                cell = ws.cell(row_i, col, g)
                if g is not None:
                    _grade_color(cell, g)
                    all_valid.append(g)
                else:
                    _style_data(cell, row_i)
                col += 1

        avg_val = round(np.mean(all_valid)) if all_valid else None
        avg_cell = ws.cell(row_i, col, avg_val)
        if avg_val is not None:
            _grade_color(avg_cell, avg_val)
        else:
            _style_data(avg_cell, row_i)

    ws.freeze_panes = "A2"
    _auto_col_widths(ws)

    ws2 = wb.create_sheet("מורים לפי מקצוע")
    ws2.append(["מקצוע", "מורה"])
    for subj, teacher in teachers.items():
        ws2.append([subj, teacher])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_events_excel_bytes(rng, np_rng, students, abilities, period, year, period_drift) -> bytes:
    wb = Workbook()
    ws = wb.active
    grade_level = students[0][3]
    ws.title = f"{grade_level} - מונה כללי"

    headers = ["מס'", "ת.ז. התלמיד", "שם התלמיד", "שכבה", "כיתה"] + [e[0] for e in EVENT_COLS]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        _style_header(cell)
    ws.row_dimensions[1].height = 20

    for row_i, (num, tz, name, grade_level, cls) in enumerate(students, 2):
        ability = abilities[row_i - 2]
        events = _generate_events_for_student(rng, np_rng, ability, period_drift=period_drift)
        for c, val in enumerate([num, tz, name, grade_level, cls], 1):
            _style_data(ws.cell(row_i, c, val), row_i)
        for c, (col_name, _, _) in enumerate(EVENT_COLS, 6):
            val = events.get(col_name)
            cell = ws.cell(row_i, c, val)
            _style_data(cell, row_i)
            if col_name == "חיסור" and val is not None and val > 40:
                cell.fill = PatternFill("solid", start_color="FFC7CE")
                cell.font = Font(name="Arial", size=9, color="9C0006")

    ws.freeze_panes = "A2"
    _auto_col_widths(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# api

def generate_school_data(
    years: list[int] | None = None,
    quarters: int = 4,
    num_students: int = 120,
    seed: int = 42,
) -> list[tuple[str, bytes, str, str, str]]:
    """
    Generate school data in memory.

    Returns a list of (filename, content_bytes, file_type, period, year) tuples
    ready to be passed directly to IngestionService.
    """
    if years is None:
        years = [2024, 2025]

    rng = random.Random(seed)
    np_rng = np.random.default_rng(seed)

    students = _generate_students(rng, num_students)
    teachers = _generate_teachers(rng, SUBJECTS)
    abilities = _student_base_ability(np_rng, num_students)
    periods = _build_periods(years, quarters)

    files: list[tuple[str, bytes, str, str, str]] = []
    for period, year, drift in periods:
        grades_bytes = _build_grades_excel_bytes(rng, np_rng, students, SUBJECTS, teachers, abilities, period, year, drift)
        events_bytes = _build_events_excel_bytes(rng, np_rng, students, abilities, period, year, drift)
        files.append((f"avg_grades_{period}_{year}.xlsx", grades_bytes, "grades", period, year))
        files.append((f"events_{period}_{year}.xlsx",     events_bytes, "events", period, year))

    return files
